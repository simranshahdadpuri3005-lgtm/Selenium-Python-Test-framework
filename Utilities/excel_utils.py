import openpyxl


def read_file(file, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, col_num).value


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    return row_count


def get_col_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    col_count = sheet.max_column
    return col_count
